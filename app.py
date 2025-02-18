import streamlit as st
st.set_page_config(
    page_title="Email PDF Analyzer",
    page_icon=":email:",
    layout="wide",
    initial_sidebar_state="collapsed"
)
# Add session state initialization at the beginning of the script
if 'page' not in st.session_state:
    st.session_state.page = 'dashboard'  # Default page is the dashboard
if 'selected_email' not in st.session_state:
    st.session_state.selected_email = None  # Track the selected email for analysis

import streamlit as st
import imaplib
import email
import os
import fitz
import base64
from datetime import datetime, timedelta
import pandas as pd
from email.header import decode_header
import pickle
import json
import random
import email.utils
import io
from PIL import Image
import zipfile
import shutil
import logging
import tempfile
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
import html

import streamlit as st
import requests
from requests.exceptions import RequestException
import pandas as pd

import requests
import re
import os
import tempfile
from urllib.parse import urlparse, parse_qs

# Add session state initialization for the new feature
if 'page' not in st.session_state:
    st.session_state.page = 'dashboard'  # Default page is the dashboard
if 'selected_email' not in st.session_state:
    st.session_state.selected_email = None  # Track the selected email for analysis

# Define the API endpoint
TESTING_CONNECT_API = 'https://pdf-analyzer-162012088916.us-central1.run.app//api/projects'


def hex_to_rgb(hex_color):
    """Convert hex color to RGB tuple."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16)/255 for i in (0, 2, 4))


def fetch_projects():
    """
    Fetch project details from the Testing Connect API.
    """
    projects = []
    error_message = None
    
    try:
        # Add headers to the request
        headers = {
            'Accept': 'application/json',
            'Origin': 'http://localhost:5001'
        }
        
        response = requests.get(TESTING_CONNECT_API, headers=headers, timeout=5, verify=False)  # Only for local development
        
        if response.status_code == 200:
            projects = response.json()
        else:
            error_message = f"Server returned status code: {response.status_code}"
    
    except RequestException as e:
        error_message = "Unable to connect to the server. Please ensure TestingConnect is running."
    except Exception as e:
        error_message = "An unexpected error occurred"
    
    return projects, error_message

def download_from_gdrive(url):
    """
    Download a file from a public Google Drive link
    
    Args:
        url (str): The Google Drive sharing URL
        
    Returns:
        tuple: (success_message, file_path) or (error_message, None)
    """
    try:
        # Extract file ID from the URL
        if 'drive.google.com' not in url:
            return "Error: Not a valid Google Drive URL", None
            
        # Extract the file ID
        if '/file/d/' in url:
            file_id = url.split('/file/d/')[1].split('/')[0]
        else:
            parsed = urlparse(url)
            file_id = parse_qs(parsed.query).get('id', [None])[0]
            
        if not file_id:
            return "Error: Could not extract file ID from URL", None
            
        # Create the direct download URL
        download_url = f"https://drive.google.com/uc?id={file_id}&export=download"
        
        # Start a session to handle the download
        session = requests.Session()
        
        # Get the initial response
        response = session.get(download_url, stream=True)
        
        # Handle large files that trigger Google's virus scan warning
        for key, value in response.cookies.items():
            if key.startswith('download_warning'):
                download_url = f"{download_url}&confirm={value}"
                response = session.get(download_url, stream=True)
                break
                
        # Get the filename from the headers if possible
        content_disposition = response.headers.get('content-disposition')
        if content_disposition:
            filename = re.findall("filename=(.+)", content_disposition)[0].strip('"')
        else:
            filename = f"gdrive_file_{file_id}.pdf"
            
        # Create a temporary file with a known path
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_filename = temp_file.name
        
        # Download the file
        with open(temp_filename, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    
        # Verify that the file was downloaded successfully
        if os.path.exists(temp_filename) and os.path.getsize(temp_filename) > 0:
            return f"Successfully downloaded: {filename}", temp_filename
        else:
            return "Error: File was not downloaded successfully", None
        
    except requests.exceptions.RequestException as e:
        return f"Error downloading file: {str(e)}", None
    except Exception as e:
        return f"An unexpected error occurred: {str(e)}", None

def apply_sidebar_styling():
    """Apply unified sidebar styling across pages"""
    st.markdown("""
        <style>
        /* Sidebar Container */
        [data-testid="stSidebar"] {
            background-color: #111827;
            padding: 1rem;
        }
        
        /* Sidebar Header */
        [data-testid="stSidebar"] .sidebar-header {
            border-bottom: 1px solid #374151;
            padding-bottom: 1rem;
            margin-bottom: 1.5rem;
        }
        
        /* Keyword Container */
        .keyword-container {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 0.75rem;
            margin: 0.5rem 0;
            border-radius: 0.5rem;
            background-color: #1F2937;
            transition: background-color 0.2s;
        }
        
        .keyword-container:hover {
            background-color: #374151;
        }
        
        /* Keyword Color Indicator */
        .keyword-color {
            width: 1rem;
            height: 1rem;
            border-radius: 0.25rem;
            margin-right: 0.75rem;
        }
        
        /* Keyword Text */
        .keyword-text {
            font-size: 0.875rem;
            font-weight: 500;
            color: #F3F4F6;
        }
        
        /* Delete Button */
        .keyword-delete-btn {
            opacity: 0;
            background: none;
            border: none;
            color: #EF4444;
            cursor: pointer;
            padding: 0.25rem;
            transition: all 0.2s;
        }
        
        .keyword-container:hover .keyword-delete-btn {
            opacity: 1;
        }
        
        .keyword-delete-btn:hover {
            color: #DC2626;
            transform: scale(1.1);
        }
        
        /* Add Keyword Section */
        .add-keyword-section {
            border-top: 1px solid #374151;
            padding-top: 1rem;
            margin-top: 1.5rem;
        }
        
        /* Input Fields */
        .stTextInput input {
            background-color: #1F2937;
            border: 1px solid #374151;
            color: #F3F4F6;
            border-radius: 0.375rem;
        }
        
        .stTextInput input:focus {
            border-color: #3B82F6;
            box-shadow: 0 0 0 2px rgba(59, 130, 246, 0.2);
        }
        
        /* Buttons */
        .stButton button {
            background-color: #3B82F6;
            color: white;
            border: none;
            border-radius: 0.375rem;
            padding: 0.5rem 1rem;
            font-weight: 500;
            width: 100%;
            transition: background-color 0.2s;
        }
        
        .stButton button:hover {
            background-color: #2563EB;
        }
        </style>
    """, unsafe_allow_html=True)


def show_keyword_sidebar(analyzer):
    """Display unified keyword management sidebar"""
    st.sidebar.markdown('<div class="sidebar-header"><h2>🔍 Keyword Management</h2></div>', unsafe_allow_html=True)
    
    # Display existing keywords
    for idx, keyword in enumerate(analyzer.KEYWORDS):
        keyword_id = sanitize_html_id(keyword)
        st.sidebar.markdown(
            f"""
            <div class="keyword-container">
                <div style="display: flex; align-items: center;">
                    <div class="keyword-color" style="background-color: {analyzer.KEYWORD_COLORS[keyword]};"></div>
                    <span class="keyword-text">{keyword}</span>
                </div>
                <button class="keyword-delete-btn" onclick="deleteKeyword('{keyword_id}_{idx}')">🗑️</button>
            </div>
            """,
            unsafe_allow_html=True
        )
    
    # Add new keyword section
    st.sidebar.markdown('<div class="add-keyword-section">', unsafe_allow_html=True)
    with st.sidebar.expander("➕ Add New Keyword", expanded=False):
        new_keyword = st.text_input("Keyword", key="new_keyword_input", placeholder="Enter keyword...")
        if new_keyword and st.button("Add Keyword", key="add_keyword_button", use_container_width=True):
            new_color = generate_color()
            analyzer.add_keyword(new_keyword, new_color)
            st.success(f"Added keyword: {new_keyword}")
            st.rerun()
    st.sidebar.markdown('</div>', unsafe_allow_html=True)

def analyze_project_document(project, analyzer):
    """
    Analyze the document from a project.
    """
    st.write(f"### Analyzing document for project: {project['name']}")
    
    try:
        with st.spinner("Downloading and analyzing document..."):
            # Download the document
            document_url = project.get('document_url')
            if not document_url:
                st.error("No document URL found for this project.")
                return
                
            # Download the file
            result, file_path = download_from_gdrive(document_url)
            
            if not file_path:
                st.error(result)
                return
                
            # Check if the file exists before trying to open it
            if not os.path.exists(file_path):
                st.error(f"File not found: {file_path}")
                return
            
            # Read and analyze the PDF
            with open(file_path, 'rb') as f:
                pdf_data = f.read()
                
            results = analyzer.analyze_pdf(pdf_data)
            
            if not results:
                st.warning("No keywords found in the document.")
                return
            
            # Display results
            for page_result in results:
                st.write(f"**Page {page_result['page_num']}**")
                
                for finding in page_result['findings']:
                    st.markdown(
                        f"""<div style='padding: 10px; 
                            background-color: {finding['color']}; 
                            margin: 5px; 
                            border-radius: 5px;'>
                            <strong>{finding['keyword']}</strong> 
                            ({finding['instances']} instances)<br>
                            Context: {finding['context']}
                        </div>""",
                        unsafe_allow_html=True
                    )
                
                # Display the highlighted page
                image = Image.open(io.BytesIO(page_result['image_data']))
                st.image(image, caption=f"Page {page_result['page_num']}", use_column_width=True)
            
            # Generate Excel report
            excel_data = generate_excel([(os.path.basename(file_path), results)])
            
            # Add download button for Excel report
            st.download_button(
                label="📥 Download Analysis Report",
                data=excel_data,
                file_name=f"analysis_report_{project['id']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Clean up the temporary file
            os.remove(file_path)
                
    except Exception as e:
        st.error(f"Error analyzing document: {str(e)}")
        logging.error(f"Error details: {str(e)}", exc_info=True)


if 'selected_project' not in st.session_state:
    st.session_state.selected_project = None

def show_testing_connect():
    """
    Display the Testing Connect view with unified styling and keyword management
    """
    if st.session_state.selected_project is None:
        show_project_list()
    else:
        show_project_analysis(st.session_state.selected_project)


def show_project_list():
    st.title("Testing Connect")
    analyzer = EmailPDFAnalyzer()
    
    # Apply unified styling
    apply_sidebar_styling()
    
    # Show keyword management sidebar
    show_keyword_sidebar(analyzer)
    
    
    # Fetch project details from the API
    projects, error_message = fetch_projects()
    
    if error_message:
        st.error(error_message)
    elif not projects:
        st.info("No projects found.")
    else:
        # Convert projects to a DataFrame and add an "Analyze" button column
        df = pd.DataFrame(projects)
        
        # Display projects in an interactive table
        st.write("### Project List")
        for idx, project in df.iterrows():
            with st.container():
                cols = st.columns([3, 1])
                with cols[0]:
                    st.markdown(f"""
                        <div style='padding: 10px; border: 1px solid #ddd; border-radius: 5px; margin: 5px 0;'>
                            <h4>{project['name']}</h4>
                            <p><strong>ID:</strong> {project['id']}</p>
                            <p><strong>Document URL:</strong> {project.get('document_url', 'N/A')}</p>
                        </div>
                    """, unsafe_allow_html=True)
                with cols[1]:
                    if st.button("📄 Analyze", key=f"analyze_{project['id']}"):
                        st.session_state.selected_project = project
                        st.rerun()

def show_project_analysis(project):
    st.title(f"Analysis Results for: {project['name']}")
    
    # Add a "Back" button to return to the project list
    if st.button("⬅️ Back to Project List"):
        st.session_state.selected_project = None
        st.rerun()
    
    analyzer = EmailPDFAnalyzer()
    analyze_project_document(project, analyzer)

def generate_excel(results):
    """
    Generate an Excel file with analysis results.
    
    Args:
        results (list): List of tuples containing (pdf_name, analysis_results)
    
    Returns:
        bytes: Excel file content as bytes
    """
    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Keep track of processed PDF names to avoid duplicates
    processed_pdfs = set()
    
    # Iterate through each PDF's results
    for pdf_name, pdf_results in results:
        # Clean up the PDF name to remove any duplicate extensions
        base_name = pdf_name.split('.')[0]
        clean_name = f"{base_name}.pdf"
        
        # Skip if we've already processed this PDF
        if clean_name in processed_pdfs:
            continue
            
        processed_pdfs.add(clean_name)
        
        # Create a valid sheet name (max 31 chars, no special chars)
        sheet_name = base_name[:31].replace('/', '_').replace('\\', '_')
        sheet = wb.create_sheet(title=sheet_name)
        
        # Add headers
        sheet.append(["Keyword", "Pages"])
        
        # Create a dictionary to store keyword-page mappings
        keyword_pages = {}
        
        # Populate the dictionary
        for page_result in pdf_results:
            page_num = page_result['page_num']
            for finding in page_result['findings']:
                keyword = finding['keyword']
                if keyword not in keyword_pages:
                    keyword_pages[keyword] = set()  # Use set to avoid duplicate page numbers
                keyword_pages[keyword].add(str(page_num))
        
        # Add data to the sheet
        for keyword, pages in keyword_pages.items():
            # Convert set to sorted list for consistent ordering
            sorted_pages = sorted(pages, key=lambda x: int(x))
            sheet.append([keyword, ", ".join(sorted_pages)])
        
        # Apply formatting
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width
        
        # Add color to the header
        for cell in sheet[1]:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Save the workbook to a bytes buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

class EmailPDFAnalyzer:
    def __init__(self):
        # Previous initialization code remains the same
        self.EMAIL = "autotake.test@gmail.com"
        self.PASSWORD = st.secrets["email_password"]
        self.DATA_FILE = "analyzed_emails.pkl"
        self.KEYWORDS_FILE = "keywords.json"
        self.load_keywords()
        self._apply_custom_styling()

    def connect_to_gmail(self):
        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(self.EMAIL, self.PASSWORD)
            return mail
        except imaplib.IMAP4.error as e:
            if "Application-specific password required" in str(e):
                st.error("""
                ### Gmail App Password Required
                
                You need to generate an App Password for Gmail:
                1. Go to Google Account settings
                2. Enable 2-Step Verification
                3. Generate App Password
                4. Add it to your .streamlit/secrets.toml file
                """)
            else:
                st.error(f"Gmail connection error: {str(e)}")
            st.stop()

    def _apply_custom_styling(self):
        """Apply custom CSS styling to unify the look across both pages"""
        st.markdown("""
        <style>
        /* Main content area */
        [data-testid="stAppViewContainer"] {
            background-color: #1E1E1E !important;
            color: #E0E0E0 !important;
            padding: 1rem !important;
        }

        /* All text in main content */
        [data-testid="stMarkdownContainer"] {
            color: #E0E0E0 !important;
        }

        /* Headers */
        h1, h2, h3, h4, h5, h6 {
            color: #FFFFFF !important;
        }

        /* Sidebar */
        [data-testid="stSidebar"] {
            background-color: #2D2D2D !important;
            border-right: 1px solid #404040 !important;
        }

        /* Sidebar text */
        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
            color: #FFFFFF !important;
        }

        /* Keyword styling in sidebar */
        [data-testid="stSidebar"] [data-testid="metric-container"] {
            background-color: transparent !important;
            padding: 5px !important;
        }

        /* Keyword label */
        [data-testid="stSidebar"] [data-testid="metric-container"] label {
            font-size: 10px !important;
            color: #FFFFFF !important;
        }

        /* Keyword value */
        [data-testid="stSidebar"] [data-testid="metric-container"] div[data-testid="stMetricValue"] {
            font-size: 5px !important;
            color: #FFFFFF !important;
        }

        /* Buttons */
        .stButton > button {
            background-color: #4A4A4A !important;
            color: #FFFFFF !important;
            border: 1px solid #606060 !important;
            border-radius: 6px !important;
            transition: all 0.3s ease !important;
        }

        .stButton > button:hover {
            background-color: #606060 !important;
            transform: scale(1.05) !important;
        }

        /* Delete button in keyword list */
        [data-testid="stSidebar"] .stButton > button {
            font-size: 12px !important;
            padding: 2px 8px !important;
            height: 24px !important;
            min-height: 24px !important;
        }

        /* Navigation radio buttons */
        .stRadio > div {
            color: #FFFFFF !important;
        }

        /* Tables */
        .stDataFrame {
            background-color: #2D2D2D !important;
            color: #E0E0E0 !important;
        }

        /* Metric containers */
        .stMetric {
            background-color: #2D2D2D !important;
            color: #FFFFFF !important;
            border: 1px solid #404040 !important;
        }

        /* Text inputs */
        .stTextInput > div > div > input {
            color: #FFFFFF !important;
            background-color: #2D2D2D !important;
            border: 1px solid #404040 !important;
            font-size: 10px !important;
        }

        /* Expandable sections */
        .streamlit-expanderHeader {
            color: #FFFFFF !important;
            background-color: #2D2D2D !important;
            font-size: 14px !important;
        }

        /* Cards and containers */
        .email-card {
            background-color: #2D2D2D !important;
            color: #E0E0E0 !important;
            border: 1px solid #404040 !important;
        }

        /* Download button */
        .stDownloadButton button {
            color: #FFFFFF !important;
            background-color: #4A4A4A !important;
        }

        /* Alerts and info boxes */
        .stAlert {
            background-color: #2D2D2D !important;
            color: #E0E0E0 !important;
        }
        </style>
        """, unsafe_allow_html=True)



    def download_pdfs_from_dropbox(self, urls, download_dir):
        from playwright.sync_api import sync_playwright
        import time
        
        downloaded_files = []
        temp_download_dir = tempfile.mkdtemp()
        
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context(
                    accept_downloads=True,
                    downloads_path=temp_download_dir
                )
                
                for url in urls:
                    try:
                        page = context.new_page()
                        page.goto(url)
                        page.wait_for_load_state('networkidle')
                        time.sleep(3)
                        
                        # Use updated XPath
                        download_button_xpath = "//*[@id='embedded-app']/span/div/div/div/div/div[2]/div/div[1]/span/div/div[2]/span[1]/button/span/span"
                        download_button = page.wait_for_selector(f"xpath={download_button_xpath}", timeout=5000)
                        
                        if download_button:
                            download_button.click()
                            time.sleep(2)
                            
                            # Handle download
                            with page.expect_download(timeout=30000) as download_info:
                                pass
                            
                            download = download_info.value
                            file_path = os.path.join(temp_download_dir, download.suggested_filename)
                            download.save_as(file_path)
                            
                            downloaded_files.append(file_path)
                            
                    except Exception as e:
                        logging.error(f"Error downloading from {url}: {str(e)}")
                    finally:
                        page.close()
                
                context.close()
                browser.close()
        
        except Exception as e:
            logging.error(f"Playwright error: {str(e)}")
        
        return downloaded_files

    def load_keywords(self):
        if os.path.exists(self.KEYWORDS_FILE):
            with open(self.KEYWORDS_FILE, 'r') as f:
                keywords_data = json.load(f)
                self.KEYWORDS = keywords_data['keywords']
                self.KEYWORD_COLORS = keywords_data['colors']
        else:
            # Initialize with default keywords
            self.KEYWORDS = ["Titus", "Price"]
            self.KEYWORD_COLORS = {
                "Titus": "#FFB6C1",  # Light pink
                "Price": "#98FB98"   # Light green
            }
            self.save_keywords()

    def save_keywords(self):
        with open(self.KEYWORDS_FILE, 'w') as f:
            json.dump({
                'keywords': self.KEYWORDS,
                'colors': self.KEYWORD_COLORS
            }, f)

    def add_keyword(self, keyword: str, color: str):
        if keyword not in self.KEYWORDS:
            self.KEYWORDS.append(keyword)
            self.KEYWORD_COLORS[keyword] = color
            self.save_keywords()

    def remove_keyword(self, keyword: str):
        if keyword in self.KEYWORDS:
            self.KEYWORDS.remove(keyword)
            del self.KEYWORD_COLORS[keyword]
            self.save_keywords()

    def parse_email_date(self, date_str):
        try:
            return email.utils.parsedate_to_datetime(date_str)
        except:
            try:
                return datetime.strptime(date_str, "%a, %d %b %Y %H:%M:%S %z")
            except:
                try:
                    return datetime.strptime(date_str.split(' (')[0], "%a, %d %b %Y %H:%M:%S %z")
                except:
                    print(f"Failed to parse date: {date_str}")
                    return datetime.now()


    def get_emails_with_pdfs(self, search_term: str = None, date_range: tuple = None):
        # Connect to Gmail and fetch emails
        mail = self.connect_to_gmail()
        mail.select("inbox")
        
        # Prepare search criteria
        search_criteria = []
        if date_range:
            start_date, end_date = date_range
            if start_date:
                search_criteria.append(f'SINCE "{start_date.strftime("%d-%b-%Y")}"')
            if end_date:
                end_date_adj = end_date + timedelta(days=1)
                search_criteria.append(f'BEFORE "{end_date_adj.strftime("%d-%b-%Y")}"')
        
        if search_term:
            search_criteria.append(f'SUBJECT "{search_term}"')
            
        if not search_criteria:
            search_criteria = ['ALL']
            
        search_string = ' '.join(search_criteria)
        _, messages = mail.search(None, search_string)
        email_list = []
        
        for num in messages[0].split():
            _, msg = mail.fetch(num, '(RFC822)')
            email_message = email.message_from_bytes(msg[0][1])
            
            # Decode subject
            subject = decode_header(email_message["subject"])[0][0]
            if isinstance(subject, bytes):
                subject = subject.decode()
            
            # Parse date
            date = self.parse_email_date(email_message["date"])
            
            # Get sender and sanitize it
            sender = email_message["from"]
            safe_sender = sanitize_email_for_display(sender)
            
            # Rest of the processing remains the same
            has_pdf = False
            dropbox_link = None
            
            # Check for PDF or Dropbox link
            for part in email_message.walk():
                if part.get_content_type() == "application/pdf":
                    has_pdf = True
                    break
                    
                if part.get_content_type() in ['text/plain', 'text/html']:
                    try:
                        body = part.get_payload(decode=True)
                        if isinstance(body, bytes):
                            body = body.decode('utf-8', errors='ignore')
                        
                        dropbox_links = re.findall(r'https?://www\.dropbox\.com/[^\s<>"]+', body)
                        if dropbox_links:
                            dropbox_link = dropbox_links[0]
                    except Exception:
                        pass
            
            # Add to email list with sanitized sender
            email_entry = {
                "id": num.decode(),
                "subject": subject,
                "date": date,
                "sender": sender,  # Original sender for display
                "sender_safe": safe_sender,  # Sanitized sender for HTML
                "analyzed": False
            }
            
            if dropbox_link:
                email_entry['dropbox_link'] = dropbox_link
            
            if has_pdf or dropbox_link:
                email_list.append(email_entry)
        
        mail.logout()
        return email_list

    def analyze_pdf(self, pdf_data):
        temp_path = "temp.pdf"
        with open(temp_path, "wb") as f:
            f.write(pdf_data)
            
        results = []
        doc = fitz.open(temp_path)
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            
            page_results = []
            found_keywords = False
            
            for keyword in self.KEYWORDS:
                if keyword.lower() in text.lower():
                    instances = page.search_for(keyword)
                    
                    if instances:
                        found_keywords = True
                        # Convert hex color to RGB tuple
                        rgb_color = hex_to_rgb(self.KEYWORD_COLORS[keyword])
                        
                        for inst in instances:
                            try:
                                # Create highlight annotation with RGB color
                                annot = page.add_highlight_annot(inst)
                                annot.set_colors(stroke=rgb_color)
                                annot.update()
                            except ValueError as e:
                                # If highlighting fails, log it but continue processing
                                logging.warning(f"Could not highlight instance of '{keyword}': {str(e)}")
                                continue
                        
                        # Enhanced context extraction
                        text_lower = text.lower()
                        keyword_lower = keyword.lower()
                        start_idx = text_lower.find(keyword_lower)
                        
                        # Extract a wider context
                        context_start = max(0, start_idx - 100)
                        context_end = min(len(text), start_idx + len(keyword) + 100)
                        context = text[context_start:context_end].strip()
                        
                        # Highlight the keyword in context
                        context_highlighted = context.replace(
                            keyword, 
                            f"**{keyword}**"
                        )
                        
                        page_results.append({
                            "keyword": keyword,
                            "color": self.KEYWORD_COLORS[keyword],
                            "instances": len(instances),
                            "context": context_highlighted
                        })
            
            if found_keywords:
                # Convert page to image with highlights
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better quality
                img_data = pix.tobytes("png")
                
                results.append({
                    "page_num": page_num + 1,
                    "findings": page_results,
                    "image_data": img_data
                })
        
        doc.close()
        os.remove(temp_path)
        return results

    def _enhance_context(self, context, window=150):
        """Enhance context extraction with intelligent trimming"""
        # Center the keyword in the context
        centered_context = context.center(window, '…')
        return centered_context.strip()


    def save_analysis_state(self, analyzed_emails):
        # Create a dictionary to track analyzed email IDs with timezone awareness
        state = {
            email_id: {
                'date': datetime.now().astimezone(),  # Include timezone info
                'subject': email_data.get('subject', 'Unknown'),
                'sender': email_data.get('sender', 'Unknown'),
                'timezone': datetime.now().astimezone().tzinfo.tzname(None)  # Store timezone name
            }
            for email_id, email_data in analyzed_emails.items()
        }
        
        try:
            with open(self.DATA_FILE, 'wb') as f:
                pickle.dump(state, f)
        except Exception as e:
            logging.error(f"Error saving analysis state: {e}")

    def load_analysis_state(self):
        try:
            if os.path.exists(self.DATA_FILE) and os.path.getsize(self.DATA_FILE) > 0:
                with open(self.DATA_FILE, 'rb') as f:
                    return pickle.load(f)
        except (EOFError, pickle.UnpicklingError, FileNotFoundError) as e:
            logging.error(f"Error loading analysis state: {e}")
        
        # Return empty dictionary if loading fails
        return {}

    def download_pdfs_from_dropbox(self, urls, download_dir):
        from playwright.sync_api import sync_playwright
        import time
        import os
        import logging
        """
        Download PDFs from multiple Dropbox links.
        
        Args:
            urls (list): List of Dropbox URLs
            download_dir (str): Directory to save downloaded PDFs
        
        Returns:
            list: Paths of downloaded PDF files
        """
        downloaded_pdfs = []
        
        for url in urls:
            try:
                # Clean and preprocess the URL
                url = url.replace('&amp;', '&')
                print(f"Processing Dropbox URL: {url}")
                
                with sync_playwright() as p:
                    logging.info("Launching browser...")
                    browser = p.chromium.launch(headless=True)
                    context = browser.new_context(accept_downloads=True)
                    
                    try:
                        page = context.new_page()
                        logging.info(f"Navigating to URL: {url}")
                        
                        # Navigate to the URL
                        page.goto(url)
                        
                        # Wait for the page to load completely
                        page.wait_for_load_state('networkidle')
                        time.sleep(3)  # Additional wait for dynamic content
                        
                        logging.info("Using specific XPath for download button...")
                        
                        # Use the provided XPath for the download button
                        download_button_xpath = "//*[@id='embedded-app']/span/div/div/div/div/div[2]/div/div[1]/span/div/div[2]/span[1]/button/span/span"
                        try:
                            download_button = page.wait_for_selector(f"xpath={download_button_xpath}", timeout=5000)
                            if download_button:
                                logging.info("Found download button, clicking...")
                                download_button.click()
                            else:
                                raise Exception("Download button not found")
                        except Exception as e:
                            logging.error(f"Failed to find or click download button: {str(e)}")
                            raise
                        
                        # Wait for the download dialog
                        time.sleep(2)
                        
                        logging.info("Looking for download-only option...")
                        
                        # Try multiple approaches to find and click the download-only option
                        download_only_selectors = [
                            "span:has-text('continue with download only')",
                            "button:has-text('continue with download only')",
                            "span.dig-Button-content[data-dig-button-content='true']",
                            "text=Or continue with download only",
                            "//span[contains(text(), 'continue with download only')]",
                            "//button[contains(text(), 'continue with download only')]"
                        ]
                        
                        download_only_clicked = False
                        for selector in download_only_selectors:
                            try:
                                if selector.startswith("//"):
                                    # Handle XPath selectors
                                    element = page.wait_for_selector(f"xpath={selector}", timeout=5000)
                                else:
                                    element = page.wait_for_selector(selector, timeout=5000)
                                
                                if element:
                                    logging.info(f"Found download-only option with selector: {selector}")
                                    # Try multiple click methods
                                    try:
                                        element.click(force=True)
                                        download_only_clicked = True
                                        break
                                    except:
                                        # Try clicking parent element if span click fails
                                        parent = element.evaluate("el => el.parentElement")
                                        if parent:
                                            page.evaluate("el => el.click()", parent)
                                            download_only_clicked = True
                                            break
                            except Exception as e:
                                logging.info(f"Selector {selector} not found or not clickable, trying next...")
                                continue
                        
                        if not download_only_clicked:
                            raise Exception("Could not find or click download-only option")
                        
                        # Wait for download to start
                        logging.info("Waiting for download to start...")
                        with page.expect_download(timeout=30000) as download_info:
                            pass
                        
                        download = download_info.value
                        
                        # Save the file
                        suggested_filename = download.suggested_filename
                        download_path = os.path.join(download_dir, suggested_filename)
                        logging.info(f"Saving file to: {download_path}")
                        download.save_as(download_path)
                        
                        logging.info(f"Download completed: {suggested_filename}")
                        downloaded_pdfs.append(download_path)
                        
                    except Exception as e:
                        logging.error(f"Dropbox download failed for {url}: {str(e)}")
                    
                    finally:
                        logging.info("Closing browser...")
                        context.close()
                        browser.close()
            
            except Exception as e:
                logging.error(f"Error processing URL {url}: {str(e)}")
        
        return downloaded_pdfs

# Modify the main method to handle multiple Dropbox links
def extract_dropbox_links(email_message):
    """
    Extract Dropbox links from an email message.
    
    Args:
        email_message (email.message.Message): Email message object
    
    Returns:
        list: List of Dropbox links found in the email
    """
    dropbox_links = []
    
    # Check email parts for Dropbox links
    for part in email_message.walk():
        if part.get_content_type() in ['text/plain', 'text/html']:
            try:
                body = part.get_payload(decode=True)
                if isinstance(body, bytes):
                    body = body.decode('utf-8', errors='ignore')
                
                # Look for Dropbox links
                import re
                found_links = re.findall(r'https?://www\.dropbox\.com/[^\s<>"\']+', body)
                dropbox_links.extend(found_links)
            except Exception as e:
                logging.error(f"Error extracting Dropbox links: {str(e)}")
    
    return dropbox_links


def generate_color():
    hue = random.random()
    saturation = 0.3 + random.random() * 0.2
    value = 0.9 + random.random() * 0.1
    
    import colorsys
    rgb = colorsys.hsv_to_rgb(hue, saturation, value)
    
    return '#{:02x}{:02x}{:02x}'.format(
        int(rgb[0] * 255),
        int(rgb[1] * 255),
        int(rgb[2] * 255)
    )
def sanitize_html_tag(tag):
    """Replace invalid characters for HTML tag names."""
    import re
    return re.sub(r'[^a-zA-Z0-9_-]', '_', tag)

def extract_and_analyze_zip(zip_path, analyzer):
    """
    Extract ZIP file and analyze all PDFs within it.
    Returns list of tuples containing (pdf_name, analysis_results)
    """
    results = []
    temp_dir = tempfile.mkdtemp()
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Walk through the extracted directory
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_path = os.path.join(root, file)
                    with open(pdf_path, 'rb') as f:
                        pdf_data = f.read()
                    analysis_results = analyzer.analyze_pdf(pdf_data)
                    results.append((file, analysis_results))
    
    finally:
        # Clean up temporary directory
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    return results

def sanitize_email_for_display(email_str):
    """
    Sanitize email address for safe display in HTML.
    Converts all special characters to underscores to ensure valid HTML IDs.
    
    Args:
        email_str (str): The email address to sanitize.
    
    Returns:
        str: A sanitized version of the email address safe for HTML IDs.
    """
    if not email_str:
        return "unknown"
    
    # Extract email from potential "Name <email@domain.com>" format
    email_match = re.search(r'<(.+?)>|(\S+@\S+)', email_str)
    if email_match:
        email_addr = email_match.group(1) or email_match.group(2)
    else:
        email_addr = email_str
    
    # Replace all non-alphanumeric characters with underscores
    safe_email = re.sub(r'[^a-zA-Z0-9]', '_', email_addr)
    
    # Ensure the ID starts with a letter (HTML requirement)
    if safe_email[0].isdigit():
        safe_email = 'e' + safe_email
        
    return safe_email

def sanitize_html_id(text):
    """
    Sanitize text for use as HTML ID.
    Ensures compliance with HTML5 ID naming rules.
    """
    if not text:
        return "unknown"
    
    # Remove any HTML tags if present
    text = re.sub(r'<[^>]+>', '', text)
    
    # Replace all non-alphanumeric characters with underscores
    safe_id = re.sub(r'[^a-zA-Z0-9]', '_', text)
    
    # Ensure ID starts with a letter (HTML requirement)
    if not safe_id[0].isalpha():
        safe_id = 'id_' + safe_id
    
    return safe_id

def display_email_card(email_data, analyzed_state):
    # Generate a completely safe unique ID for this email card
    card_id = sanitize_html_id(f"email_{email_data.get('id', '')}_{email_data.get('sender', '')}")
    
    # Format date properly
    date_str = email_data['date'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(email_data['date'], datetime) else str(email_data['date'])
    
    # Check if this email has been analyzed (more robust check)
    is_analyzed = False
    if email_data['id'] in analyzed_state:
        analyzed_info = analyzed_state[email_data['id']]
        is_analyzed = (
            analyzed_info.get('subject') == email_data.get('subject', '') and
            analyzed_info.get('sender') == email_data.get('sender', '')
        )
    
    # Escape HTML special characters in display text
    safe_subject = html.escape(str(email_data.get('subject', '')))
    safe_sender = html.escape(str(email_data.get('sender', '')))
    
    with st.container():
        st.markdown(
            f"""
            <div id="{card_id}" class="email-card" style="
                border: 1px solid #ddd;
                border-radius: 10px;
                padding: 15px;
                margin: 10px 0;
                background-color: {'#e6f7ff' if is_analyzed else '#ffffff'};
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            ">
                <h4>{safe_subject}</h4>
                <p><strong>From:</strong> {safe_sender}</p>
                <p><strong>Date:</strong> {date_str}</p>
                <p><strong>Status:</strong> {'Analyzed ✅' if is_analyzed else 'Not Analyzed ❌'}</p>
            </div>
            """,
            unsafe_allow_html=True
        )
        
        # Use sanitized ID for button key
        button_key = f"analyze_{card_id}"
        if st.button("Analyze PDF", key=button_key):
            return True
            
    return False


# Add this function to handle email filtering
def filter_emails(df, exclude_domains=None):
    """
    Filter out emails from specific domains with proper error handling.
    """
    if exclude_domains is None:
        exclude_domains = ['bookmyshow.com']
    
    if df.empty:
        return df
    
    if 'sender' not in df.columns:
        return df
    
    # Create a safe mask for filtering
    mask = df['sender'].notna()
    for domain in exclude_domains:
        # Safely convert to string and check for domain
        mask &= ~df['sender'].astype(str).str.lower().str.contains(
            re.escape(domain), 
            na=False, 
            regex=True
        )
    
    return df[mask]



# Modify the main function to handle page navigation
def main():
    # Initialize session state if not already done
    if 'page' not in st.session_state:
        st.session_state.page = 'dashboard'
    if 'selected_email' not in st.session_state:
        st.session_state.selected_email = None
    
    # Navigation in sidebar
    st.sidebar.title("Navigation")
    page = st.sidebar.radio("Go to", ["Email PDF Analyzer", "Testing Connect"])
    
    # Update session state based on navigation
    if page == "Email PDF Analyzer":
        if st.session_state.page not in ['dashboard', 'email_analysis']:
            st.session_state.page = 'dashboard'
    elif page == "Testing Connect":
        st.session_state.page = 'testing_connect'
    
    # Display the appropriate page
    if st.session_state.page == 'dashboard':
        show_dashboard()
    elif st.session_state.page == 'email_analysis':
        show_email_analysis()
    elif st.session_state.page == 'testing_connect':
        show_testing_connect()
    elif st.session_state.page == 'project_analysis':
        # Keep existing TestingConnect analysis page
        show_project_analysis(st.session_state.selected_project)


# Add this function to display the dashboard
def show_dashboard():
    st.title("📄 Email PDF Analyzer")
    analyzer = EmailPDFAnalyzer()
    
    # Apply unified styling
    apply_sidebar_styling()
    
    # Show keyword management sidebar
    show_keyword_sidebar(analyzer)
    
    st.write("### Search and Filter")
    col1, col2, col3 = st.columns([3, 1, 1])
    
    with col1:
        search_term = st.text_input("🔎 Search Email Subject", placeholder="Enter search term...")
    
    with col2:
        start_date = st.date_input("Start Date", datetime.now() - timedelta(days=7))
    
    with col3:
        end_date = st.date_input("End Date", datetime.now())
    
    if start_date > end_date:
        st.error("Error: Start date must be before end date")
        return
        
    date_range = (start_date, end_date) if start_date and end_date else None
    
    analyzed_state = analyzer.load_analysis_state()
    emails = analyzer.get_emails_with_pdfs(search_term, date_range)
    
    if not emails:
        st.info("No emails found matching the search criteria.")
        return
    
    # Convert emails to DataFrame
    df = pd.DataFrame(emails)
    
    # Filter out unwanted emails
    df = filter_emails(df, exclude_domains=['bookmyshow.com'])
    
    # Display emails in a grid layout
    st.write("### Email List")
    cols = st.columns(2)
    for idx, email_data in enumerate(df.to_dict('records')):
        with cols[idx % 2]:
            if display_email_card(email_data, analyzed_state):
                st.session_state.selected_email = email_data
                st.session_state.page = 'email_analysis'  # New page state for email analysis
                st.rerun()

def show_email_analysis():
    st.title("📄 Email Analysis Results")
    
    # Add a "Back" button to return to the dashboard
    if st.button("⬅️ Back to Email Dashboard"):
        st.session_state.page = 'dashboard'
        st.session_state.selected_email = None
        st.rerun()
    
    # Get the selected email and analyze it
    email_data = st.session_state.selected_email
    if email_data:
        analyzer = EmailPDFAnalyzer()
        analyzed_state = analyzer.load_analysis_state()
        analyze_email(email_data, analyzer, analyzed_state)
    else:
        st.error("No email selected for analysis.")
        if st.button("Return to Dashboard"):
            st.session_state.page = 'dashboard'
            st.rerun()

# Add this function to display the results page
def show_results():
    st.title("📄 Analysis Results")
    
    # Add a "Back" button to return to the dashboard
    if st.button("⬅️ Back to Dashboard"):
        st.session_state.page = 'dashboard'
        st.session_state.selected_email = None
        st.rerun()
    
    # Analyze the selected email
    analyzer = EmailPDFAnalyzer()
    analyzed_state = analyzer.load_analysis_state()
    email_data = st.session_state.selected_email
    
    if email_data:
        analyze_email(email_data, analyzer, analyzed_state)
    else:
        st.error("No email selected for analysis.")


# Modify the analyze_email function to display results on the new page
def analyze_email(email_data, analyzer, analyzed_state):
    st.title(f"Analysis Results for: {email_data['subject']}")
    
    with st.spinner("Analyzing PDF..."):
        try:
            mail = analyzer.connect_to_gmail()
            mail.select("inbox")
            _, msg = mail.fetch(email_data['id'].encode(), '(RFC822)')
            email_message = email.message_from_bytes(msg[0][1])
            
            dropbox_links = extract_dropbox_links(email_message)
            all_results = []

            if dropbox_links:
                # Create temporary directory for downloads
                temp_download_dir = tempfile.mkdtemp()
                try:
                    downloaded_files = analyzer.download_pdfs_from_dropbox(dropbox_links, temp_download_dir)
                    
                    for file_path in downloaded_files:
                        if file_path.lower().endswith('.zip'):
                            # Handle ZIP files
                            pdf_results = extract_and_analyze_zip(file_path, analyzer)
                            for pdf_name, results in pdf_results:
                                st.subheader(f"Analysis for {pdf_name}")
                                all_results.append((pdf_name, results))
                                
                                # Display results for this PDF
                                for page_result in results:
                                    st.write(f"**Page {page_result['page_num']}**")
                                    
                                    for finding in page_result['findings']:
                                        st.markdown(
                                            f"<div style='padding: 10px; background-color: {finding['color']}; margin: 5px;'>"
                                            f"<strong>{finding['keyword']}</strong> "
                                            f"({finding['instances']} instances)<br>"
                                            f"Context: {finding['context']}"
                                            f"</div>",
                                            unsafe_allow_html=True
                                        )
                                    
                                    image = Image.open(io.BytesIO(page_result['image_data']))
                                    st.image(image, caption=f"Page {page_result['page_num']}", use_column_width=True) 
                                                           
                        elif file_path.lower().endswith('.pdf'):
                            # Handle individual PDFs
                            with open(file_path, 'rb') as f:
                                pdf_data = f.read()
                            results = analyzer.analyze_pdf(pdf_data)
                            all_results.append((os.path.basename(file_path), results))
                
                finally:
                    # Clean up all temporary files and directories
                    shutil.rmtree(temp_download_dir, ignore_errors=True)
            
            else:
                # Handle PDFs directly attached to email
                for part in email_message.walk():
                    if part.get_content_type() == "application/pdf":
                        # Get the original filename from the attachment
                        filename = part.get_filename()
                        if not filename:
                            # If filename not found in Content-Disposition header, try Content-Type header
                            filename = part.get_param('name')
                        
                        # If still no filename, try to decode the header
                        if not filename and 'Content-Disposition' in part:
                            try:
                                disposition = email.header.decode_header(part['Content-Disposition'])
                                for content, charset in disposition:
                                    if isinstance(content, bytes):
                                        content = content.decode(charset or 'utf-8')
                                    if 'filename=' in content:
                                        filename = content.split('filename=')[1].strip('"\'')
                                        break
                            except:
                                pass
                        
                        # If still no filename, use a generic name with timestamp
                        if not filename:
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f"attached_pdf_{timestamp}.pdf"
                        
                        # Clean the filename
                        filename = re.sub(r'[<>:"/\\|?*]', '_', filename)  # Remove invalid characters
                        if not filename.lower().endswith('.pdf'):
                            filename += '.pdf'
                        
                        pdf_data = base64.b64decode(part.get_payload())
                        results = analyzer.analyze_pdf(pdf_data)
                        all_results.append((filename, results))

                        # Display results for this PDF
                        st.subheader(f"Analysis for {filename}")
                        for page_result in results:
                            st.write(f"**Page {page_result['page_num']}**")
                            
                            for finding in page_result['findings']:
                                st.markdown(
                                    f"<div style='padding: 10px; background-color: {finding['color']}; margin: 5px;'>"
                                    f"<strong>{finding['keyword']}</strong> "
                                    f"({finding['instances']} instances)<br>"
                                    f"Context: {finding['context']}"
                                    f"</div>",
                                    unsafe_allow_html=True
                                )
                            
                            image = Image.open(io.BytesIO(page_result['image_data']))
                            st.image(image, caption=f"Page {page_result['page_num']}", use_column_width=True)

            if not all_results:
                st.error("No PDFs found in the email or Dropbox links.")
                return

            # Update analysis state
            analyzed_state[email_data['id']] = {
                'date': datetime.now().astimezone(),  # Include timezone
                'subject': email_data.get('subject', 'Unknown'),
                'sender': email_data.get('sender', 'Unknown'),
                'timezone': datetime.now().astimezone().tzinfo.tzname(None)
            }
            analyzer.save_analysis_state(analyzed_state)
            st.success("Analysis complete!")

            # Generate Excel file
            excel_data = generate_excel(all_results)
            
            # Add download button
            st.download_button(
                label="📥 Download Results",
                data=excel_data,
                file_name="analysis_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            mail.logout()

        except Exception as e:
            st.error(f"Error processing PDFs: {str(e)}")
            logging.error(f"Error details: {str(e)}", exc_info=True)


if __name__ == "__main__":
    main()









    